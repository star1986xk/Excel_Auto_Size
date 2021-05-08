import os
import sys
import random
from copy import deepcopy
from ui.ui_main import Ui_MainWindow
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
import openpyxl

from mylogclass import MyLogClass

SIZE = ('Small', 'Medium', 'Large', 'X-Large', 'XX-Large', '3X-Large', '4X-Large', '5X-Large', '6X-Large')


class MainWindow(QMainWindow, Ui_MainWindow):

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.retranslateUi(self)
        self.setFixedSize(self.width(), self.height())

        self.log = MyLogClass()

        self.signals()

    def signals(self):
        self.pushButton_excel.clicked.connect(lambda: self.open_file(self.lineEdit_excel, 'Excel (*.xlsx)'))
        self.pushButton_run.clicked.connect(self.run)

    # 槽函数-----------------------------------------------------------------------------------------------------------
    # 打开文件
    def open_file(self, lineEdit, file_type):
        try:
            filename, _ = QFileDialog.getOpenFileName(self, '选取文件', './', file_type)
            lineEdit.setText(filename)
        except Exception as e:
            self.log.logger.warning(str(e))

    # 生成主逻辑
    def run(self):
        try:
            path_excel = self.lineEdit_excel.text()
            if not path_excel: return

            self.wb = openpyxl.load_workbook(path_excel)
            self.ws = self.wb.active

            start_row, start_column = self.position()
            self.set_excel(start_row, start_column)
            self.wb.save(path_excel)
            QMessageBox.information(self, '提示', '运行完成')
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
            self.log.logger.warning(str(e))

    def position(self):
        row_count = self.ws.max_row
        column_count = self.ws.max_column

        for r in range(1, row_count + 1):
            for c in range(1, column_count + 1):
                if self.ws.cell(r, c).value == 'Alpha':
                    return r, c

    def set_excel(self, start_row, start_column):
        for n, li in enumerate(SIZE):
            self.ws.cell(start_row + n, start_column + 1).value = li
        if self.ws.cell(start_row + 10, start_column).value == 'Alpha':
            return self.set_excel(start_row + 10, start_column)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
